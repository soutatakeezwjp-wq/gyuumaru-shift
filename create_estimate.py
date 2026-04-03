#!/usr/bin/env python3
"""
合同会社kaleido - ぎゅう丸シフト管理システム見積書作成スクリプト
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from datetime import date

def create_estimate():
    wb = openpyxl.Workbook()

    # --- スタイル定義 ---
    # 罫線
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='medium'))
    double_bottom = Border(bottom=Side(style='double'))

    # フォント
    title_font = Font(name='Yu Gothic', size=18)
    header_font = Font(name='Yu Gothic', size=11)
    item_font = Font(name='Yu Gothic', size=10)
    total_font = Font(name='Yu Gothic', size=14)
    small_font = Font(name='Yu Gothic', size=9)
    note_font = Font(name='Yu Gothic', size=9, color='666666')

    # 塗りつぶし
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font_white = Font(name='Yu Gothic', size=10, color='FFFFFF')
    light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    total_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    grand_total_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    grand_total_font = Font(name='Yu Gothic', size=16, color='FFFFFF')
    subtotal_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')

    # アラインメント
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # --- 3パターンのデータ定義 ---
    plans = [
        {
            'name': 'ライトプラン',
            'subtitle': '基本機能のみ / コストを抑えたい方向け',
            'initial_items': [
                {
                    'category': '初期開発費用',
                    'items': [
                        ('システム設計・要件定義', '画面構成・DB設計・API仕様策定', 1, 150000),
                        ('フロントエンド開発', 'スタッフ向け3画面 + 管理者向け5画面のUI構築', 1, 250000),
                        ('バックエンド開発（API・DB構築）', 'Cloudflare Workers + D1 / 全APIエンドポイント実装', 1, 200000),
                        ('シフト自動作成エンジン開発', '連勤・インターバル・週上限チェック付き配置アルゴリズム', 1, 200000),
                        ('テスト・デプロイ・導入支援', '動作検証・本番環境構築・初期データ投入・操作説明', 1, 100000),
                    ]
                }
            ],
            'monthly_items': [
                {
                    'category': '月額運用費用',
                    'items': [
                        ('サーバー・インフラ利用料', 'Cloudflare Workers / D1 / 独自ドメイン維持費', 1, 5000),
                        ('システム保守・監視', '稼働監視・セキュリティアップデート・バグ修正', 1, 15000),
                        ('メール・チャットサポート', '営業時間内の問い合わせ対応（月5件まで）', 1, 10000),
                    ]
                }
            ]
        },
        {
            'name': 'スタンダードプラン',
            'subtitle': '全機能搭載 / 多店舗運用におすすめ（推奨）',
            'initial_items': [
                {
                    'category': '初期開発費用',
                    'items': [
                        ('システム設計・要件定義', '画面構成・DB設計・API仕様策定・運用フロー設計', 1, 250000),
                        ('フロントエンド開発', 'スタッフ向け3画面 + 管理者向け8画面 / レスポンシブ対応', 1, 400000),
                        ('バックエンド開発（API・DB構築）', 'Cloudflare Workers + D1 / 28エンドポイント / JWT認証', 1, 350000),
                        ('シフト自動作成エンジン開発', 'ポジション別・曜日別最低人数 / 正社員労働時間管理', 1, 300000),
                        ('LINE通知・Excel出力機能', 'LINE Messaging API連携 / シフト表Excel出力', 1, 200000),
                        ('テスト・デプロイ・導入支援', '全店舗テスト・本番環境構築・マニュアル作成・操作研修', 1, 200000),
                    ]
                }
            ],
            'monthly_items': [
                {
                    'category': '月額運用費用',
                    'items': [
                        ('サーバー・インフラ利用料', 'Cloudflare Workers / D1 / 独自ドメイン / SSL証明書', 1, 5000),
                        ('システム保守・監視', '24時間稼働監視・セキュリティ対応・バグ修正・性能最適化', 1, 25000),
                        ('サポート対応', '営業時間内の問い合わせ対応（月10件まで）/ リモート対応', 1, 15000),
                        ('軽微な機能改善', '月2件までの小規模な画面修正・機能追加対応', 1, 15000),
                    ]
                }
            ]
        },
        {
            'name': 'プレミアムプラン',
            'subtitle': '専任サポート付き / カスタマイズ・拡張も全てお任せ',
            'initial_items': [
                {
                    'category': '初期開発費用',
                    'items': [
                        ('システム設計・要件定義', '業務ヒアリング・画面設計・DB設計・API仕様・運用設計', 1, 400000),
                        ('フロントエンド開発', '全13画面 / レスポンシブ / アニメーション / UX最適化', 1, 600000),
                        ('バックエンド開発（API・DB構築）', 'Cloudflare Workers + D1 / 28+ API / JWT認証 / ログ基盤', 1, 500000),
                        ('シフト自動作成エンジン開発', 'AI最適化 / ポジション・曜日・繁忙期対応 / 学習型配置', 1, 500000),
                        ('LINE通知・Excel出力・外部連携', 'LINE API / Excel出力 / 勤怠システム連携 / Google Calendar連携', 1, 350000),
                        ('テスト・デプロイ・導入支援', '全店舗テスト・本番構築・マニュアル・研修（現地2回）', 1, 350000),
                    ]
                }
            ],
            'monthly_items': [
                {
                    'category': '月額運用費用',
                    'items': [
                        ('サーバー・インフラ利用料', 'Cloudflare Workers Pro / D1 / 独自ドメイン / 高可用性構成', 1, 10000),
                        ('システム保守・監視・障害対応', '24時間監視 / 緊急時2時間以内対応 / セキュリティ監査', 1, 40000),
                        ('専任サポート', '専任担当者 / 電話・メール・チャット無制限 / 月次報告', 1, 30000),
                        ('機能改善・追加開発', '月4件までの機能追加・改善 / 優先対応', 1, 30000),
                        ('データバックアップ・DR対策', '日次バックアップ / 災害復旧計画 / データ移行サポート', 1, 10000),
                    ]
                }
            ]
        }
    ]

    today = date.today()
    estimate_date = today.strftime('%Y年%m月%d日')
    valid_until = f'{today.year}年{today.month + 1 if today.month < 12 else 1:02d}月{today.day:02d}日'
    estimate_no = f'KLD-{today.strftime("%Y%m%d")}-001'

    # --- シート作成 ---
    for idx, plan in enumerate(plans):
        if idx == 0:
            ws = wb.active
            ws.title = plan['name']
        else:
            ws = wb.create_sheet(title=plan['name'])

        # 列幅設定
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20

        # 印刷設定
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'portrait'
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        row = 2

        # --- ヘッダー部分 ---
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '御 見 積 書'
        cell.font = title_font
        cell.alignment = center
        row += 2

        # 見積番号・日付
        ws[f'E{row}'].value = '見積番号:'
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = right_align
        ws[f'F{row}'].value = estimate_no
        ws[f'F{row}'].font = small_font
        ws[f'F{row}'].alignment = right_align
        row += 1

        ws[f'E{row}'].value = '見積日:'
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = right_align
        ws[f'F{row}'].value = estimate_date
        ws[f'F{row}'].font = small_font
        ws[f'F{row}'].alignment = right_align
        row += 1

        ws[f'E{row}'].value = '有効期限:'
        ws[f'E{row}'].font = small_font
        ws[f'E{row}'].alignment = right_align
        ws[f'F{row}'].value = valid_until
        ws[f'F{row}'].font = small_font
        ws[f'F{row}'].alignment = right_align
        row += 2

        # 宛先
        ws.merge_cells(f'B{row}:C{row}')
        cell = ws[f'B{row}']
        cell.value = 'ぎゅう丸 御中'
        cell.font = Font(name='Yu Gothic', size=14)
        cell.border = Border(bottom=Side(style='medium'))
        ws[f'C{row}'].border = Border(bottom=Side(style='medium'))
        row += 2

        # プラン名
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = f'【{plan["name"]}】{plan["subtitle"]}'
        cell.font = Font(name='Yu Gothic', size=11, color='2C3E50')
        cell.alignment = left_align
        row += 1

        # 件名
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '件名: シフト管理システム（ぎゅう丸専用）開発・導入・運用'
        cell.font = Font(name='Yu Gothic', size=10)
        row += 2

        # ==================== 初期費用セクション ====================
        # セクションタイトル
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '■ 初期費用（税別）'
        cell.font = Font(name='Yu Gothic', size=12)
        row += 1

        # ヘッダー行
        headers = ['項目', '内容', '数量', '単価（円）', '金額（円）']
        cols = ['B', 'C', 'D', 'E', 'F']
        for c, h in zip(cols, headers):
            cell = ws[f'{c}{row}']
            cell.value = h
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        # 初期費用アイテム
        initial_total = 0
        for cat in plan['initial_items']:
            for i, (name, desc, qty, price) in enumerate(cat['items']):
                amount = qty * price
                initial_total += amount

                bg = light_gray_fill if i % 2 == 0 else None

                cell_b = ws[f'B{row}']
                cell_b.value = name
                cell_b.font = item_font
                cell_b.alignment = wrap
                cell_b.border = thin_border
                if bg:
                    cell_b.fill = bg

                cell_c = ws[f'C{row}']
                cell_c.value = desc
                cell_c.font = small_font
                cell_c.alignment = wrap
                cell_c.border = thin_border
                if bg:
                    cell_c.fill = bg

                cell_d = ws[f'D{row}']
                cell_d.value = qty
                cell_d.font = item_font
                cell_d.alignment = center
                cell_d.border = thin_border
                if bg:
                    cell_d.fill = bg

                cell_e = ws[f'E{row}']
                cell_e.value = price
                cell_e.font = item_font
                cell_e.alignment = right_align
                cell_e.number_format = '#,##0'
                cell_e.border = thin_border
                if bg:
                    cell_e.fill = bg

                cell_f = ws[f'F{row}']
                cell_f.value = amount
                cell_f.font = item_font
                cell_f.alignment = right_align
                cell_f.number_format = '#,##0'
                cell_f.border = thin_border
                if bg:
                    cell_f.fill = bg

                ws.row_dimensions[row].height = 30
                row += 1

        # 初期費用小計
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '初期費用 小計（税別）'
        cell.font = Font(name='Yu Gothic', size=11)
        cell.alignment = right_align
        cell.border = thin_border
        cell.fill = subtotal_fill
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
            ws[f'{c}{row}'].fill = subtotal_fill

        cell_f = ws[f'F{row}']
        cell_f.value = initial_total
        cell_f.font = Font(name='Yu Gothic', size=12)
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        cell_f.fill = subtotal_fill
        row += 1

        # 消費税
        tax = int(initial_total * 0.1)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '消費税（10%）'
        cell.font = item_font
        cell.alignment = right_align
        cell.border = thin_border
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border

        cell_f = ws[f'F{row}']
        cell_f.value = tax
        cell_f.font = item_font
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        row += 1

        # 初期費用合計（税込）
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '初期費用 合計（税込）'
        cell.font = Font(name='Yu Gothic', size=12, color='FFFFFF')
        cell.alignment = right_align
        cell.border = thin_border
        cell.fill = grand_total_fill
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
            ws[f'{c}{row}'].fill = grand_total_fill

        cell_f = ws[f'F{row}']
        cell_f.value = initial_total + tax
        cell_f.font = grand_total_font
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        cell_f.fill = grand_total_fill
        ws.row_dimensions[row].height = 35
        row += 2

        # ==================== 月額費用セクション ====================
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '■ 月額運用費用（税別）'
        cell.font = Font(name='Yu Gothic', size=12)
        row += 1

        # ヘッダー行
        for c, h in zip(cols, headers):
            cell = ws[f'{c}{row}']
            cell.value = h
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        row += 1

        # 月額費用アイテム
        monthly_total = 0
        for cat in plan['monthly_items']:
            for i, (name, desc, qty, price) in enumerate(cat['items']):
                amount = qty * price
                monthly_total += amount

                bg = light_gray_fill if i % 2 == 0 else None

                cell_b = ws[f'B{row}']
                cell_b.value = name
                cell_b.font = item_font
                cell_b.alignment = wrap
                cell_b.border = thin_border
                if bg:
                    cell_b.fill = bg

                cell_c = ws[f'C{row}']
                cell_c.value = desc
                cell_c.font = small_font
                cell_c.alignment = wrap
                cell_c.border = thin_border
                if bg:
                    cell_c.fill = bg

                cell_d = ws[f'D{row}']
                cell_d.value = qty
                cell_d.font = item_font
                cell_d.alignment = center
                cell_d.border = thin_border
                if bg:
                    cell_d.fill = bg

                cell_e = ws[f'E{row}']
                cell_e.value = price
                cell_e.font = item_font
                cell_e.alignment = right_align
                cell_e.number_format = '#,##0'
                cell_e.border = thin_border
                if bg:
                    cell_e.fill = bg

                cell_f = ws[f'F{row}']
                cell_f.value = amount
                cell_f.font = item_font
                cell_f.alignment = right_align
                cell_f.number_format = '#,##0'
                cell_f.border = thin_border
                if bg:
                    cell_f.fill = bg

                ws.row_dimensions[row].height = 30
                row += 1

        # 月額小計
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '月額費用 小計（税別）'
        cell.font = Font(name='Yu Gothic', size=11)
        cell.alignment = right_align
        cell.border = thin_border
        cell.fill = subtotal_fill
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
            ws[f'{c}{row}'].fill = subtotal_fill

        cell_f = ws[f'F{row}']
        cell_f.value = monthly_total
        cell_f.font = Font(name='Yu Gothic', size=12)
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        cell_f.fill = subtotal_fill
        row += 1

        # 消費税
        monthly_tax = int(monthly_total * 0.1)
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '消費税（10%）'
        cell.font = item_font
        cell.alignment = right_align
        cell.border = thin_border
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border

        cell_f = ws[f'F{row}']
        cell_f.value = monthly_tax
        cell_f.font = item_font
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        row += 1

        # 月額合計（税込）
        ws.merge_cells(f'B{row}:E{row}')
        cell = ws[f'B{row}']
        cell.value = '月額費用 合計（税込）'
        cell.font = Font(name='Yu Gothic', size=12, color='FFFFFF')
        cell.alignment = right_align
        cell.border = thin_border
        cell.fill = grand_total_fill
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
            ws[f'{c}{row}'].fill = grand_total_fill

        cell_f = ws[f'F{row}']
        cell_f.value = monthly_total + monthly_tax
        cell_f.font = grand_total_font
        cell_f.alignment = right_align
        cell_f.number_format = '#,##0'
        cell_f.border = thin_border
        cell_f.fill = grand_total_fill
        ws.row_dimensions[row].height = 35
        row += 2

        # ==================== 年間コスト試算 ====================
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '■ 年間コスト試算（税込・参考）'
        cell.font = Font(name='Yu Gothic', size=12)
        row += 1

        annual_monthly = (monthly_total + monthly_tax) * 12
        annual_total = (initial_total + tax) + annual_monthly

        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = '初年度合計（初期費用 + 月額12ヶ月分）'
        ws[f'B{row}'].font = item_font
        ws[f'B{row}'].alignment = right_align
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].fill = total_fill
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
            ws[f'{c}{row}'].fill = total_fill
        ws[f'F{row}'].value = annual_total
        ws[f'F{row}'].font = Font(name='Yu Gothic', size=13)
        ws[f'F{row}'].alignment = right_align
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].fill = total_fill
        ws.row_dimensions[row].height = 30
        row += 1

        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].value = '2年目以降（月額のみ / 年間）'
        ws[f'B{row}'].font = item_font
        ws[f'B{row}'].alignment = right_align
        ws[f'B{row}'].border = thin_border
        for c in ['C', 'D', 'E']:
            ws[f'{c}{row}'].border = thin_border
        ws[f'F{row}'].value = annual_monthly
        ws[f'F{row}'].font = Font(name='Yu Gothic', size=11)
        ws[f'F{row}'].alignment = right_align
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'F{row}'].border = thin_border
        row += 2

        # ==================== 備考 ====================
        ws.merge_cells(f'B{row}:F{row}')
        cell = ws[f'B{row}']
        cell.value = '■ 備考'
        cell.font = Font(name='Yu Gothic', size=11)
        row += 1

        notes = [
            '・上記金額は全て税別表記です（消費税10%は別途記載）。',
            '・見積有効期限は発行日より1ヶ月間です。',
            '・初期費用のお支払いは、着手時50%・納品時50%の2回払いとなります。',
            '・月額費用は毎月末締め翌月末払いとなります。',
            '・開発期間の目安: 契約後 約4〜6週間で本番リリース。',
            '・対象店舗数: 13店舗（マルチテナント対応済み）。',
            '・カスタマイズや追加機能については別途お見積もりいたします。',
        ]
        for note in notes:
            ws.merge_cells(f'B{row}:F{row}')
            ws[f'B{row}'].value = note
            ws[f'B{row}'].font = note_font
            ws[f'B{row}'].alignment = wrap
            row += 1

        row += 2

        # ==================== 発行元情報 ====================
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'].value = '合同会社 kaleido'
        ws[f'D{row}'].font = Font(name='Yu Gothic', size=12)
        ws[f'D{row}'].alignment = right_align
        row += 1
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'].value = '担当: '
        ws[f'D{row}'].font = small_font
        ws[f'D{row}'].alignment = right_align
        row += 1
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'].value = 'TEL: '
        ws[f'D{row}'].font = small_font
        ws[f'D{row}'].alignment = right_align
        row += 1
        ws.merge_cells(f'D{row}:F{row}')
        ws[f'D{row}'].value = 'Email: '
        ws[f'D{row}'].font = small_font
        ws[f'D{row}'].alignment = right_align

    # --- 比較サマリーシート ---
    ws_summary = wb.create_sheet(title='プラン比較', index=0)
    ws_summary.column_dimensions['A'].width = 3
    ws_summary.column_dimensions['B'].width = 28
    ws_summary.column_dimensions['C'].width = 22
    ws_summary.column_dimensions['D'].width = 22
    ws_summary.column_dimensions['E'].width = 22

    row = 2
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'].value = '御見積書 - プラン比較表'
    ws_summary[f'B{row}'].font = title_font
    ws_summary[f'B{row}'].alignment = center
    row += 1

    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'].value = 'ぎゅう丸シフト管理システム 開発・導入・運用'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=10, color='666666')
    ws_summary[f'B{row}'].alignment = center
    row += 2

    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'].value = f'見積番号: {estimate_no}　/　見積日: {estimate_date}　/　合同会社 kaleido'
    ws_summary[f'B{row}'].font = small_font
    ws_summary[f'B{row}'].alignment = center
    row += 2

    # 比較表ヘッダー
    plan_fills = [
        PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid'),  # ライト=青
        PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid'),  # スタンダード=オレンジ
        PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid'),  # プレミアム=紫
    ]

    ws_summary[f'B{row}'].value = ''
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'B{row}'].border = thin_border

    for i, (plan, col) in enumerate(zip(plans, ['C', 'D', 'E'])):
        cell = ws_summary[f'{col}{row}']
        cell.value = plan['name']
        cell.font = Font(name='Yu Gothic', size=11, color='FFFFFF')
        cell.fill = plan_fills[i]
        cell.alignment = center
        cell.border = thin_border
    row += 1

    # 初期費用
    ws_summary[f'B{row}'].value = '初期費用（税別）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=10)
    ws_summary[f'B{row}'].border = thin_border
    ws_summary[f'B{row}'].fill = light_gray_fill

    initial_totals = []
    for plan in plans:
        total = sum(price * qty for cat in plan['initial_items'] for name, desc, qty, price in cat['items'])
        initial_totals.append(total)

    for total, col in zip(initial_totals, ['C', 'D', 'E']):
        cell = ws_summary[f'{col}{row}']
        cell.value = total
        cell.font = Font(name='Yu Gothic', size=11)
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
        cell.fill = light_gray_fill
    row += 1

    # 初期費用（税込）
    ws_summary[f'B{row}'].value = '初期費用（税込）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=11)
    ws_summary[f'B{row}'].border = thin_border

    for total, col in zip(initial_totals, ['C', 'D', 'E']):
        cell = ws_summary[f'{col}{row}']
        cell.value = int(total * 1.1)
        cell.font = Font(name='Yu Gothic', size=12)
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
    ws_summary.row_dimensions[row].height = 28
    row += 1

    # 月額費用
    ws_summary[f'B{row}'].value = '月額費用（税別）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=10)
    ws_summary[f'B{row}'].border = thin_border
    ws_summary[f'B{row}'].fill = light_gray_fill

    monthly_totals = []
    for plan in plans:
        total = sum(price * qty for cat in plan['monthly_items'] for name, desc, qty, price in cat['items'])
        monthly_totals.append(total)

    for total, col in zip(monthly_totals, ['C', 'D', 'E']):
        cell = ws_summary[f'{col}{row}']
        cell.value = total
        cell.font = Font(name='Yu Gothic', size=11)
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
        cell.fill = light_gray_fill
    row += 1

    # 月額（税込）
    ws_summary[f'B{row}'].value = '月額費用（税込）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=11)
    ws_summary[f'B{row}'].border = thin_border

    for total, col in zip(monthly_totals, ['C', 'D', 'E']):
        cell = ws_summary[f'{col}{row}']
        cell.value = int(total * 1.1)
        cell.font = Font(name='Yu Gothic', size=12)
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
    ws_summary.row_dimensions[row].height = 28
    row += 1

    # 初年度合計
    ws_summary[f'B{row}'].value = '初年度 合計（税込）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=11, color='FFFFFF')
    ws_summary[f'B{row}'].fill = grand_total_fill
    ws_summary[f'B{row}'].border = thin_border
    ws_summary[f'B{row}'].alignment = left_align

    for itotal, mtotal, col in zip(initial_totals, monthly_totals, ['C', 'D', 'E']):
        annual = int(itotal * 1.1) + int(mtotal * 1.1) * 12
        cell = ws_summary[f'{col}{row}']
        cell.value = annual
        cell.font = Font(name='Yu Gothic', size=14, color='FFFFFF')
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
        cell.fill = grand_total_fill
    ws_summary.row_dimensions[row].height = 35
    row += 1

    # 2年目以降
    ws_summary[f'B{row}'].value = '2年目以降（年間・税込）'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=10)
    ws_summary[f'B{row}'].border = thin_border
    ws_summary[f'B{row}'].fill = light_gray_fill

    for mtotal, col in zip(monthly_totals, ['C', 'D', 'E']):
        cell = ws_summary[f'{col}{row}']
        cell.value = int(mtotal * 1.1) * 12
        cell.font = Font(name='Yu Gothic', size=11)
        cell.alignment = right_align
        cell.number_format = '#,##0"円"'
        cell.border = thin_border
        cell.fill = light_gray_fill
    row += 2

    # 機能比較
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'].value = '■ 機能・サービス比較'
    ws_summary[f'B{row}'].font = Font(name='Yu Gothic', size=12)
    row += 1

    features = [
        ('スタッフ向けシフト希望入力', 'O', 'O', 'O'),
        ('管理者向けダッシュボード', 'O', 'O', 'O'),
        ('シフト自動作成エンジン', '基本', 'ポジション対応', 'AI最適化'),
        ('正社員 労働時間管理', '-', 'O', 'O'),
        ('LINE通知', '-', 'O', 'O'),
        ('Excel出力', '-', 'O', 'O'),
        ('外部システム連携', '-', '-', 'O'),
        ('サポート', 'メール(月5件)', 'メール+リモート(月10件)', '専任・無制限'),
        ('機能改善対応', '-', '月2件', '月4件'),
        ('障害対応', '翌営業日', '当日', '2時間以内'),
    ]

    # ヘッダー
    ws_summary[f'B{row}'].value = '機能・サービス'
    ws_summary[f'B{row}'].font = header_font_white
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'B{row}'].border = thin_border
    ws_summary[f'B{row}'].alignment = center
    for plan, col, fill in zip(plans, ['C', 'D', 'E'], plan_fills):
        cell = ws_summary[f'{col}{row}']
        cell.value = plan['name']
        cell.font = Font(name='Yu Gothic', size=10, color='FFFFFF')
        cell.fill = fill
        cell.alignment = center
        cell.border = thin_border
    row += 1

    for i, (feat, l, s, p) in enumerate(features):
        bg = light_gray_fill if i % 2 == 0 else None

        ws_summary[f'B{row}'].value = feat
        ws_summary[f'B{row}'].font = item_font
        ws_summary[f'B{row}'].border = thin_border
        ws_summary[f'B{row}'].alignment = left_align
        if bg:
            ws_summary[f'B{row}'].fill = bg

        for val, col in zip([l, s, p], ['C', 'D', 'E']):
            cell = ws_summary[f'{col}{row}']
            cell.value = val
            cell.font = item_font
            cell.alignment = center
            cell.border = thin_border
            if bg:
                cell.fill = bg
            if val == '-':
                cell.font = Font(name='Yu Gothic', size=10, color='CCCCCC')
        row += 1

    row += 1
    ws_summary.merge_cells(f'B{row}:E{row}')
    ws_summary[f'B{row}'].value = '※ 各プランの詳細は個別シートをご参照ください。'
    ws_summary[f'B{row}'].font = note_font
    ws_summary[f'B{row}'].alignment = center

    # 保存
    filepath = '/Users/takebayashisouta/Downloads/claude code2/ぎゅう丸アプリ開発/見積書_ぎゅう丸シフト管理システム.xlsx'
    wb.save(filepath)
    print(f'見積書を作成しました: {filepath}')

if __name__ == '__main__':
    create_estimate()
