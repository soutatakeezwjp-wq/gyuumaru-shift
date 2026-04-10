"""
ぎゅう丸シフト管理システム - ROIシミュレーター & 見積書Excel生成
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_roi_excel():
    wb = openpyxl.Workbook()

    # 色定義
    BROWN = "4A3323"
    CREAM = "FFF8E1"
    WHITE = "FFFFFF"
    LIGHT_BROWN = "F5F0EB"
    GREEN = "E8F5E9"
    GREEN_TXT = "2E7D32"
    RED = "FFEBEE"
    RED_TXT = "C62828"
    BLUE = "E3F2FD"
    BLUE_TXT = "1565C0"
    ORANGE = "FFF3E0"
    ORANGE_TXT = "E65100"
    GRAY = "F5F5F5"
    BORDER_COLOR = "D4C5B5"

    thin = Side(style='thin', color=BORDER_COLOR)
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick_bottom = Border(top=thin, bottom=Side(style='medium', color=BROWN), left=thin, right=thin)

    # ===== Sheet 1: ROIシミュレーター =====
    ws1 = wb.active
    ws1.title = "ROIシミュレーター"
    ws1.sheet_properties.tabColor = GREEN_TXT

    # 列幅
    for col, w in [(1,4),(2,30),(3,20),(4,20),(5,20),(6,20),(7,20)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    # タイトル
    ws1.merge_cells('B1:F1')
    c = ws1['B1']
    c.value = "ぎゅう丸シフト管理システム ROIシミュレーター"
    c.font = Font(name='Yu Gothic', bold=True, size=18, color=BROWN)
    c.fill = PatternFill('solid', fgColor=CREAM)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 50

    ws1.merge_cells('B2:F2')
    ws1['B2'].value = "導入前後でどれだけコスト削減できるかを試算します"
    ws1['B2'].font = Font(name='Yu Gothic', size=11, color="888888")
    ws1['B2'].alignment = Alignment(horizontal='center')

    # 入力セクション
    row = 4
    ws1.merge_cells(f'B{row}:F{row}')
    c = ws1[f'B{row}']
    c.value = "入力項目（黄色のセルに数値を入力してください）"
    c.font = Font(name='Yu Gothic', bold=True, size=14, color=WHITE)
    c.fill = PatternFill('solid', fgColor=BROWN)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[row].height = 35

    inputs = [
        ("店舗数", 13, "店舗", "ぎゅう丸の全店舗数"),
        ("1店舗あたりのスタッフ数", 20, "名", "正社員+アルバイト合計"),
        ("店長の平均時給換算", 2500, "円/時", "月給÷160時間で算出"),
        ("シフト作成にかかる時間（現在）", 8, "時間/月", "1店舗あたり、毎月の作成時間"),
        ("シフト調整・連絡にかかる時間（現在）", 4, "時間/月", "電話・LINE等での調整時間"),
        ("シフト作成にかかる時間（導入後）", 1, "時間/月", "自動作成+微調整のみ"),
        ("シフト調整・連絡にかかる時間（導入後）", 0.5, "時間/月", "アプリ内で完結"),
    ]

    for i, (label, default, unit, note) in enumerate(inputs):
        r = row + 1 + i
        ws1[f'B{r}'].value = label
        ws1[f'B{r}'].font = Font(name='Yu Gothic', size=11, color=BROWN)
        ws1[f'B{r}'].border = border

        ws1[f'C{r}'].value = default
        ws1[f'C{r}'].font = Font(name='Yu Gothic', bold=True, size=14, color=BROWN)
        ws1[f'C{r}'].fill = PatternFill('solid', fgColor="FFFFCC")
        ws1[f'C{r}'].alignment = Alignment(horizontal='center')
        ws1[f'C{r}'].border = border
        ws1[f'C{r}'].number_format = '#,##0.0'

        ws1[f'D{r}'].value = unit
        ws1[f'D{r}'].font = Font(name='Yu Gothic', size=10, color="888888")
        ws1[f'D{r}'].border = border

        ws1[f'E{r}'].value = note
        ws1[f'E{r}'].font = Font(name='Yu Gothic', size=10, color="888888")
        ws1[f'E{r}'].border = border

    # 計算結果セクション
    calc_row = row + len(inputs) + 2
    ws1.merge_cells(f'B{calc_row}:F{calc_row}')
    c = ws1[f'B{calc_row}']
    c.value = "削減効果（自動計算）"
    c.font = Font(name='Yu Gothic', bold=True, size=14, color=WHITE)
    c.fill = PatternFill('solid', fgColor=GREEN_TXT)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[calc_row].height = 35

    # 数式で自動計算
    stores = f'C{row+1}'
    staff = f'C{row+2}'
    hourly = f'C{row+3}'
    before_create = f'C{row+4}'
    before_adjust = f'C{row+5}'
    after_create = f'C{row+6}'
    after_adjust = f'C{row+7}'

    calcs = [
        ("月間シフト作成コスト（現在）", f'={stores}*{hourly}*{before_create}', "円/月"),
        ("月間シフト調整コスト（現在）", f'={stores}*{hourly}*{before_adjust}', "円/月"),
        ("月間合計コスト（現在）", f'=C{calc_row+1}+C{calc_row+2}', "円/月"),
        ("", "", ""),
        ("月間シフト作成コスト（導入後）", f'={stores}*{hourly}*{after_create}', "円/月"),
        ("月間シフト調整コスト（導入後）", f'={stores}*{hourly}*{after_adjust}', "円/月"),
        ("月間合計コスト（導入後）", f'=C{calc_row+5}+C{calc_row+6}', "円/月"),
        ("", "", ""),
        ("月間削減額", f'=C{calc_row+3}-C{calc_row+7}', "円/月"),
        ("年間削減額", f'=C{calc_row+9}*12', "円/年"),
        ("", "", ""),
        ("初期費用", 1500000, "円"),
        ("月額費用", 80000, "円/月"),
        ("年間ランニングコスト", f'=C{calc_row+13}*12', "円/年"),
        ("", "", ""),
        ("年間純削減額（削減額-ランニングコスト）", f'=C{calc_row+10}-C{calc_row+14}', "円/年"),
        ("投資回収期間", f'=IF(C{calc_row+9}-C{calc_row+13}>0,C{calc_row+12}/(C{calc_row+9}-C{calc_row+13}),"削減効果なし")', "ヶ月"),
        ("3年間のROI", f'=IF(C{calc_row+12}>0,(C{calc_row+16}*3-C{calc_row+12})/C{calc_row+12}*100,"N/A")', "%"),
    ]

    for i, (label, formula, unit) in enumerate(calcs):
        r = calc_row + 1 + i
        if not label:
            continue
        ws1[f'B{r}'].value = label
        ws1[f'B{r}'].font = Font(name='Yu Gothic', size=11, color=BROWN)
        ws1[f'B{r}'].border = border

        ws1[f'C{r}'].value = formula
        ws1[f'C{r}'].number_format = '#,##0'
        ws1[f'C{r}'].border = border

        ws1[f'D{r}'].value = unit
        ws1[f'D{r}'].font = Font(name='Yu Gothic', size=10, color="888888")
        ws1[f'D{r}'].border = border

        # ハイライト
        if "削減額" in label or "ROI" in label or "回収" in label:
            ws1[f'B{r}'].font = Font(name='Yu Gothic', bold=True, size=12, color=GREEN_TXT)
            ws1[f'C{r}'].font = Font(name='Yu Gothic', bold=True, size=16, color=GREEN_TXT)
            ws1[f'C{r}'].fill = PatternFill('solid', fgColor=GREEN)
        elif "初期費用" in label or "月額費用" in label:
            ws1[f'C{r}'].font = Font(name='Yu Gothic', bold=True, size=14, color=BLUE_TXT)
            ws1[f'C{r}'].fill = PatternFill('solid', fgColor=BLUE)

    # ===== Sheet 2: 競合比較 =====
    ws2 = wb.create_sheet("競合比較")
    ws2.sheet_properties.tabColor = BLUE_TXT

    for col, w in [(1,4),(2,22),(3,18),(4,18),(5,18),(6,18),(7,18),(8,22)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.merge_cells('B1:H1')
    c = ws2['B1']
    c.value = "シフト管理システム 競合比較表"
    c.font = Font(name='Yu Gothic', bold=True, size=18, color=BROWN)
    c.fill = PatternFill('solid', fgColor=CREAM)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 50

    # ヘッダー
    headers = ["サービス名", "初期費用", "月額費用\n(13店舗想定)", "主な機能", "カスタマイズ", "開発期間", "備考"]
    for i, h in enumerate(headers):
        c = ws2.cell(row=3, column=i+2, value=h)
        c.font = Font(name='Yu Gothic', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', fgColor=BROWN)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws2.row_dimensions[3].height = 40

    competitors = [
        ["ぎゅう丸専用\n(kaleido開発)", "150万円", "8万円/月\n(年96万円)", "シフト自動作成\n時間帯別人数管理\n余剰自動解消\n正社員時間管理\nガントチャート\nExcel出力", "完全カスタム\n自由に対応可能", "開発済み\n即導入可", "月額0円の競合と比較して\n圧倒的なカスタマイズ性\nと専用設計が強み"],
        ["Airシフト\n(リクルート)", "0円", "約5.7万円/月\n(330円x20人x13店)", "シフト収集\nシフト作成\nチャット機能", "不可\n(汎用サービス)", "-", "汎用的だが\nカスタマイズ不可"],
        ["ジョブカン", "0円", "約10.4万円/月\n(400円x20人x13店)", "シフト管理\n勤怠管理\n自動作成", "不可", "-", "勤怠込みだが\n飲食特化ではない"],
        ["シフオプ\n(リクルート)", "0円", "約8.6万円/月\n(330円x20人x13店)", "ガントチャート\n人件費概算\n人員過不足表示", "不可", "-", "ガントチャートが\n強みだが固定仕様"],
        ["らくしふ", "要問合せ\n(数十万円)", "要問合せ\n(月10-20万円)", "LINE連携\n自動リマインド\nヘルプ募集", "一部対応\n(有料)", "1-2ヶ月", "LINE連携が強み\nだがコスト高"],
        ["アールシフト", "0円", "10万円~/月", "AI自動作成\n800種類の機能\nスキル考慮", "一部対応\n(有料)", "1-3ヶ月", "大規模向け\n中小には過剰"],
        ["oplus", "0円", "約1.3万円/月\n(100円x20人x13店)", "基本シフト管理\nメガホン機能", "不可", "-", "最安だが\n機能は基本的"],
        ["スクラッチ開発\n(他社)", "300-500万円", "5-15万円/月\n(保守費)", "要件次第", "完全カスタム", "3-6ヶ月", "開発期間とコストが\n大きなリスク"],
    ]

    for i, row_data in enumerate(competitors):
        r = 4 + i
        ws2.row_dimensions[r].height = 80
        for j, val in enumerate(row_data):
            c = ws2.cell(row=r, column=j+2, value=val)
            c.font = Font(name='Yu Gothic', size=10, color=BROWN)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

            if i == 0:  # ぎゅう丸行をハイライト
                c.fill = PatternFill('solid', fgColor=GREEN)
                c.font = Font(name='Yu Gothic', bold=True, size=10, color=GREEN_TXT)

    # 3年間トータルコスト比較
    cost_row = 4 + len(competitors) + 2
    ws2.merge_cells(f'B{cost_row}:H{cost_row}')
    c = ws2[f'B{cost_row}']
    c.value = "3年間トータルコスト比較（13店舗・スタッフ20名/店の場合）"
    c.font = Font(name='Yu Gothic', bold=True, size=14, color=WHITE)
    c.fill = PatternFill('solid', fgColor=BROWN)
    c.alignment = Alignment(horizontal='center')
    ws2.row_dimensions[cost_row].height = 35

    cost_headers = ["サービス", "初期費用", "月額(税込)", "年間費用", "3年間合計", "ぎゅう丸との差"]
    for i, h in enumerate(cost_headers):
        c = ws2.cell(row=cost_row+1, column=i+2, value=h)
        c.font = Font(name='Yu Gothic', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', fgColor="8D6E63")
        c.alignment = Alignment(horizontal='center')
        c.border = border

    cost_data = [
        ["ぎゅう丸専用", 1500000, 80000, 960000, 1500000+960000*3, 0],
        ["Airシフト", 0, 57200, 686400, 686400*3, None],
        ["ジョブカン", 0, 104000, 1248000, 1248000*3, None],
        ["シフオプ", 0, 85800, 1029600, 1029600*3, None],
        ["らくしふ", 500000, 150000, 1800000, 500000+1800000*3, None],
        ["アールシフト", 0, 100000, 1200000, 1200000*3, None],
        ["スクラッチ開発", 4000000, 100000, 1200000, 4000000+1200000*3, None],
    ]

    gyumaru_3y = cost_data[0][4]
    for i, d in enumerate(cost_data):
        r = cost_row + 2 + i
        if d[5] is None:
            d[5] = d[4] - gyumaru_3y

        for j, val in enumerate(d):
            c = ws2.cell(row=r, column=j+2, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center')
            if j == 0:
                c.font = Font(name='Yu Gothic', bold=True, size=10, color=BROWN)
            elif j >= 1:
                c.number_format = '#,##0'
                c.font = Font(name='Yu Gothic', size=10)

            if i == 0:  # ぎゅう丸行
                c.fill = PatternFill('solid', fgColor=GREEN)
                if j >= 1:
                    c.font = Font(name='Yu Gothic', bold=True, size=11, color=GREEN_TXT)

            # 差額が正(ぎゅう丸が安い)なら赤表示
            if j == 5 and isinstance(val, (int, float)):
                if val > 0:
                    c.font = Font(name='Yu Gothic', bold=True, size=11, color=RED_TXT)
                    c.value = f"+{val:,.0f}円"
                elif val < 0:
                    c.font = Font(name='Yu Gothic', bold=True, size=11, color=GREEN_TXT)
                    c.value = f"{val:,.0f}円"
                else:
                    c.value = "-"

    # ===== Sheet 3: 機能比較 =====
    ws3 = wb.create_sheet("機能比較")
    ws3.sheet_properties.tabColor = ORANGE_TXT

    for col, w in [(1,4),(2,28),(3,12),(4,12),(5,12),(6,12),(7,12),(8,12)]:
        ws3.column_dimensions[get_column_letter(col)].width = w

    ws3.merge_cells('B1:H1')
    c = ws3['B1']
    c.value = "機能比較表"
    c.font = Font(name='Yu Gothic', bold=True, size=18, color=BROWN)
    c.fill = PatternFill('solid', fgColor=CREAM)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 50

    func_headers = ["機能", "ぎゅう丸\n専用", "Airシフト", "ジョブカン", "シフオプ", "らくしふ", "oplus"]
    for i, h in enumerate(func_headers):
        c = ws3.cell(row=3, column=i+2, value=h)
        c.font = Font(name='Yu Gothic', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', fgColor=BROWN)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws3.row_dimensions[3].height = 40

    OK = "○"
    NG = "×"
    TRI = "△"

    features = [
        ["シフト希望収集", OK, OK, OK, OK, OK, OK],
        ["シフト自動作成", OK, TRI, OK, NG, TRI, NG],
        ["時間帯別必要人数管理", OK, OK, TRI, OK, TRI, NG],
        ["余剰自動解消", OK, NG, NG, NG, NG, NG],
        ["ガントチャート表示", OK, TRI, NG, OK, NG, NG],
        ["ドラッグ&ドロップ編集", OK, OK, NG, NG, NG, NG],
        ["正社員 月間時間管理", OK, NG, OK, NG, NG, NG],
        ["人件費自動計算", OK, TRI, OK, NG, NG, NG],
        ["ステップ型ワークフロー", OK, NG, NG, NG, NG, NG],
        ["不足文面自動生成", OK, NG, NG, NG, NG, NG],
        ["Excel出力", OK, NG, OK, NG, TRI, NG],
        ["LINE連携", TRI, NG, OK, NG, OK, NG],
        ["複数店舗管理", OK, OK, OK, OK, OK, OK],
        ["カスタマイズ対応", OK, NG, NG, NG, TRI, NG],
        ["モバイル対応", OK, OK, OK, OK, OK, OK],
        ["専用設計（飲食特化）", OK, NG, NG, NG, NG, NG],
    ]

    for i, row_data in enumerate(features):
        r = 4 + i
        for j, val in enumerate(row_data):
            c = ws3.cell(row=r, column=j+2, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
            if j == 0:
                c.font = Font(name='Yu Gothic', size=11, color=BROWN)
                c.alignment = Alignment(horizontal='left', vertical='center')
            elif val == OK:
                c.font = Font(name='Yu Gothic', bold=True, size=14, color=GREEN_TXT)
            elif val == NG:
                c.font = Font(name='Yu Gothic', size=14, color="CCCCCC")
            elif val == TRI:
                c.font = Font(name='Yu Gothic', size=14, color=ORANGE_TXT)

            # ぎゅう丸列をハイライト
            if j == 1:
                c.fill = PatternFill('solid', fgColor=GREEN)

    # 保存
    filepath = "/Users/takebayashisouta/Projects/福岡AI/gyumaru-app/ROI_シミュレーター_ぎゅう丸シフト管理.xlsx"
    wb.save(filepath)
    print(f"保存完了: {filepath}")
    return filepath

if __name__ == "__main__":
    create_roi_excel()
